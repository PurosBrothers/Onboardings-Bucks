#!/usr/bin/env python3
"""
Script para procesar el archivo de modelo de causación y almacenar los datos en MongoDB en la coleccion client_pucs.
"""
import os
import sys
import logging
from datetime import datetime
from pathlib import Path
from bson import ObjectId
from dotenv import load_dotenv
import openpyxl
from config.mongodb_config import MongoDBConfig
from utils.mongodb_manager import MongoDBManager
from config.beanie_config import init_db

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Cargar variables de entorno
load_dotenv()

def get_headers_from_row(wb, row_number=5):
    """
    Obtiene los encabezados del archivo Excel a partir de la fila especificada.
    
    Args:
        wb: Workbook de openpyxl
        row_number: Número de fila donde están los encabezados (por defecto 5)
        
    Returns:
        list: Lista de encabezados
    """
    ws = wb["Hoja1"]  # Usar "Hoja1"
    headers = []
    
    # Obtener la fila de encabezados
    header_row = list(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))[0]
    
    # Filtrar valores None y convertir a string
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    
    logger.info(f"📋 Encabezados encontrados: {headers}")
    return headers

def create_client_pucs_indexes(uid_user: str):
    """
    Crea los índices necesarios para la colección client_pucs.
    
    Args:
        uid_user: ID del usuario desde el archivo .env
    """
    try:
        # Inicializar configuración de MongoDB
        config = MongoDBConfig(env_prefix="DEV")
        config.set_collection_name("client_pucs")
        
        # Inicializar MongoDB Manager
        manager = MongoDBManager(config)
        
        # Crear índice compuesto único para UID y cuenta_contable
        manager.collection.create_index(
            [("UID", 1), ("cuenta_contable", 1)], 
            unique=True,
            name="uid_cuenta_contable_unique"
        )
        
        logger.info("🔗 Índice compuesto único creado: UID + cuenta_contable")
        
        # Cerrar conexión a MongoDB
        manager.close()
        
    except Exception as e:
        logger.error(f"❌ Error creando índices: {str(e)}")
        # No exit here - indexes might already exist
        logger.warning("⚠️ Continuando sin crear índices...")

def delete_existing_client_pucs(uid_user: str):
    """
    Elimina todos los documentos existentes para el usuario en la colección client_pucs.
    
    Args:
        uid_user: ID del usuario desde el archivo .env
    """
    try:
        # Inicializar configuración de MongoDB
        config = MongoDBConfig(env_prefix="DEV")
        config.set_collection_name("client_pucs")
        
        # Inicializar MongoDB Manager
        manager = MongoDBManager(config)
        
        # Eliminar todos los documentos para el usuario
        result = manager.collection.delete_many({"UID": ObjectId(uid_user)})
        
        logger.info(f"🗑️ Eliminados {result.deleted_count} documentos existentes para el usuario {uid_user}")
        
        # Cerrar conexión a MongoDB
        manager.close()
        
    except Exception as e:
        logger.error(f"❌ Error eliminando documentos existentes: {str(e)}")
        sys.exit(1)

def process_excel_file(uid_user: str, xlsx_path: str):
    """
    Procesa el archivo Excel y almacena los datos en MongoDB.
    
    Args:
        uid_user: ID del usuario desde el archivo .env
        xlsx_path: Ruta absoluta al archivo Excel
    """
    try:
        # Verificar que el archivo existe
        if not os.path.exists(xlsx_path):
            logger.error(f"❌ Archivo Excel no encontrado: {xlsx_path}")
            sys.exit(1)
        
        # Eliminar documentos existentes para el usuario
        logger.info("🧹 Eliminando documentos existentes...")
        delete_existing_client_pucs(uid_user)
        
        # Crear índices necesarios
        logger.info("🔗 Creando índices...")
        create_client_pucs_indexes(uid_user)
            
        # Inicializar configuración de MongoDB
        config = MongoDBConfig(env_prefix="DEV")
        config.set_collection_name("client_pucs")
        
        # Inicializar MongoDB Manager
        manager = MongoDBManager(config)
        
        # Cargar el archivo Excel
        logger.info(f"📄 Cargando archivo Excel: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        
        # Verificar que "Hoja1" existe
        if "Hoja1" not in wb.sheetnames:
            logger.error(f"❌ Hoja 'Hoja1' no encontrada en el archivo. Hojas disponibles: {wb.sheetnames}")
            sys.exit(1)
        
        # Obtener encabezados
        headers = get_headers_from_row(wb)
        
        # Obtener la hoja "Hoja1"
        ws = wb["Hoja1"]
        
        # Procesar filas a partir de la fila 6 (después de los encabezados)
        row_count = 0
        skipped_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
            # Obtener el valor de la columna "CUENTA CONTABLE   (OBLIGATORIO)" (columna D)
            cuenta_contable_idx = headers.index("CUENTA CONTABLE   (OBLIGATORIO)")
            cuenta_contable = str(row[cuenta_contable_idx]).strip() if row[cuenta_contable_idx] is not None else ""
            centro_costo_idx = headers.index("CENTRO DE COSTO")
            centro_costo = str(row[centro_costo_idx]).strip() if row[centro_costo_idx] is not None else ""
            nit_idx = headers.index("NIT")
            nit = str(row[nit_idx]).strip() if row[nit_idx] is not None else ""
            centro_costo_sub_idx = headers.index("SUBCENTRO DE COSTO")
            centro_costo_sub = str(row[centro_costo_sub_idx]).strip() if row[centro_costo_sub_idx] is not None else ""
            
            # Si la cuenta contable está vacía, saltar esta fila
            if not cuenta_contable:
                continue
            
            # Verificar si la cuenta contable ya existe para este usuario
            existing_doc = manager.collection.find_one({
                "UID": ObjectId(uid_user),
                "cuenta_contable": cuenta_contable
            })
            
            if existing_doc:
                logger.warning(f"⚠️ Cuenta contable '{cuenta_contable}' ya existe, saltando...")
                skipped_count += 1
                continue
            
            puc = cuenta_contable[:6] 
            
            if search_embedding(puc): 
            # Crear documento para MongoDB
                document = {
                    "UID": ObjectId(uid_user),
                    "cuenta_contable": cuenta_contable,
                    "description": "",  # Campo vacío por ahora
                    "code_field": "",   # Campo vacío por ahora
                    "createdAt": datetime.now(),
                    "updatedAt": datetime.now()
                }
            
            else:
                document = {
                    "UID": ObjectId(uid_user),
                    "cuenta_contable": cuenta_contable,
                    "description": "uniquePuc",  # Campo vacío por ahora
                    "code_field": "",   # Campo vacío por ahora
                    "createdAt": datetime.now(),
                    "updatedAt": datetime.now()
                }
            
            # Insertar documento en MongoDB
            manager.collection.insert_one(document)
            create_cost_center_per_puc(uid_user,nit,cuenta_contable,centro_costo, centro_costo_sub)
            row_count += 1

            # Log el documento creado
            logger.info(f"📝 Documento creado: {document}")
            
            # Log cada 100 filas procesadas
            if row_count % 100 == 0:
                logger.info(f"✅ Procesadas {row_count} filas...")
        
        logger.info(f"🎉 Proceso completado. Se procesaron {row_count} filas en total.")
        if skipped_count > 0:
            logger.info(f"⚠️ Se saltaron {skipped_count} filas con cuentas contables duplicadas.")
        
        # Cerrar conexión a MongoDB
        manager.close()
        
    except Exception as e:
        logger.error(f"❌ Error procesando archivo Excel: {str(e)}")
        sys.exit(1)
        
def search_embedding (puc:str):
    try:
        config = MongoDBConfig(env_prefix="DEV")
        config.set_collection_name("puc_embeddings")
        manager_embedding = MongoDBManager(config)
        existing_puc = manager_embedding.collection.find_one({"code": puc})
        if existing_puc:
            return True
        else:
            return False
    except Exception as e:
        logger.error(f"❌ Error buscando embedding para el PUC {puc}: {str(e)}")
        return False


def create_cost_center_per_puc(uid: str, nit: str, account_code: str, cost_center: str, centro_costo_sub: str = None):
    """
    Crea un documento de centro de costo por PUC en la colección correspondiente.
    Los nombres de los atributos se envían en inglés y se agregan logs detallados en español.
    """
    logger.info(f"Iniciando creación de centro de costo para PUC: UID={uid}, NIT={nit}, Cuenta={account_code}, CentroCosto={cost_center}, SubCentro={centro_costo_sub}")
    
    # Inicializar configuración de MongoDB
    config = MongoDBConfig(env_prefix="DEV")
    config.set_collection_name("cost_center_per_puc")
    managers = MongoDBManager(config)
    nit_clean = limpiar_nit(nit)
    
    # Construir el objeto cost_center como JSON
    cost_center_obj = {
        "center": cost_center
    }
    
    # Agregar subcentro si existe
    if centro_costo_sub:
        cost_center_obj["subcenter"] = centro_costo_sub
    
    # Buscar si ya existe
    existing_doc = managers.collection.find_one({
        "uid": ObjectId(uid),
        "id_supplier": nit_clean,
        "account_code": account_code,
        "cost_center": cost_center_obj,
    })
    
    if existing_doc:
        logger.warning(f"⚠️ Ya existe un centro de costo para la cuenta '{account_code}' y NIT '{nit_clean}', se omite...")
        return
    
    document = {
        "uid": ObjectId(uid),
        "id_supplier": nit_clean,
        "account_code": account_code,
        "cost_center": cost_center_obj
    }
    
    result = managers.collection.insert_one(document)
    logger.info(f"✅ Centro de costo por PUC creado con _id: {result.inserted_id} para NIT: {nit_clean}, Cuenta: {account_code}, CentroCosto: {cost_center_obj}")
        
    
def limpiar_nit(nit):
    if isinstance(nit, str):
        return ''.join(filter(str.isdigit, nit))
    return nit

def main():
    """
    Función principal que ejecuta el proceso completo.
    """
    logger.info("🚀 Iniciando procesamiento del archivo de causación...")
    
    # Obtener UID_USER del archivo .env
    uid_user = os.getenv("UID_USER")
    if not uid_user:
        logger.error("❌ Variable de entorno UID_USER no encontrada en el archivo .env")
        sys.exit(1)
    
    # Obtener la ruta absoluta del archivo Excel
    app_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx_path = os.path.abspath(os.path.join(app_root, "data", "modelos_causacion", "SurtifloraModeloCausacionAbril2025.xlsx"))
    
    # Procesar el archivo
    process_excel_file(uid_user, xlsx_path)
    logger.info("✨ Proceso completado exitosamente.")

if __name__ == "__main__":
    main()
