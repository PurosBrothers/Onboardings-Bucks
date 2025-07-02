#!/usr/bin/env python3
"""
Script para procesar el archivo de modelo de causación y almacenar los datos en MongoDB en la colección client_pucs.
"""

# =============================
# Imports y configuración global
# =============================
import os
import sys
import logging
from datetime import datetime
from pathlib import Path
from bson import ObjectId
import openpyxl

from config.mongodb_config import MongoDBConfig
from utils.mongodb_manager import MongoDBManager
from config.beanie_config import init_db


# =============================
# Configuración de logging
# =============================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# =============================
# Utilidades de Excel
# =============================
def obtener_encabezados_excel(libro, fila_encabezados=5):
    """
    Obtiene los encabezados del archivo Excel a partir de la fila especificada.
    """
    hoja = libro["Hoja1"]
    fila = list(hoja.iter_rows(min_row=fila_encabezados, max_row=fila_encabezados, values_only=True))[0]
    encabezados = [str(h).strip() if h is not None else "" for h in fila]
    logger.info(f"Encabezados encontrados: {encabezados}")
    return encabezados

def obtener_code_field(libro, cuenta_contable):
    try:
        if "Hoja5" not in libro.sheetnames:
            logger.warning("Hoja5 no encontrada en el archivo Excel")
            return ["default_code_field"]  # Devolver como array
        
        hoja = libro["Hoja5"]
        cuenta_contable = str(cuenta_contable).strip()  # Normalizar a cadena
        logger.info(f"Buscando code_field para cuenta_contable: {cuenta_contable}")

        # Verificar las filas leídas de Hoja5
        logger.debug("Iniciando iteración sobre filas de Hoja5...")
        puc_values = []
        for idx, fila in enumerate(hoja.iter_rows(min_row=5, values_only=True), start=5):
            logger.debug(f"Fila {idx}: {fila}")  # Registrar todos los valores de la fila
            if fila[1] is not None:  # Usar columna B (índice 1) para los códigos PUC
                codigo_puc = str(fila[1]).strip()
                puc_values.append(codigo_puc)
                logger.debug(f"Leyendo PUC desde columna B (fila {idx}, col 1): {codigo_puc} (tipo: {type(fila[1])})")
        
        logger.debug(f"Valores de PUC encontrados en Hoja5: {puc_values}")

        # Buscar coincidencia exacta
        for fila in hoja.iter_rows(min_row=5, values_only=True):
            if fila[1] is not None:  # Usar columna B (índice 1) para los códigos PUC
                codigo_puc = str(fila[1]).strip()
                logger.debug(f"Comparando PUC '{codigo_puc}' (tipo: {type(codigo_puc)}) con cuenta_contable '{cuenta_contable}' (tipo: {type(cuenta_contable)})")
                if codigo_puc == cuenta_contable:
                    code_field_raw = fila[4] if fila[4] is not None else None  # Columna E (índice 4)
                    
                    # Procesar el code_field como array
                    if code_field_raw is not None:
                        code_field_str = str(code_field_raw).strip()
                        # Verificar si hay saltos de línea y dividir
                        if '\n' in code_field_str:
                            code_field_array = [item.strip() for item in code_field_str.split('\n') if item.strip()]
                        else:
                            code_field_array = [code_field_str] if code_field_str else ["default_code_field"]
                    else:
                        code_field_array = ["default_code_field"]
                    
                    logger.info(f"Coincidencia exacta encontrada. Raw code_field (col E): {code_field_raw}, Procesado como array: {code_field_array}")
                    return code_field_array
        
        # Fallback: buscar por los primeros 6 dígitos
        puc_6_digitos = cuenta_contable[:6]
        logger.info(f"Fallback: Buscando con los primeros 6 dígitos: {puc_6_digitos}")
        for fila in hoja.iter_rows(min_row=5, values_only=True):
            if fila[1] is not None:
                codigo_puc = str(fila[1]).strip()
                logger.debug(f"Fallback: Comparando PUC '{codigo_puc}' con '{puc_6_digitos}'")
                if codigo_puc.startswith(puc_6_digitos):
                    code_field_raw = fila[4] if fila[4] is not None else None  # Columna E (índice 4)
                    
                    # Procesar el code_field como array
                    if code_field_raw is not None:
                        code_field_str = str(code_field_raw).strip()
                        # Verificar si hay saltos de línea y dividir
                        if '\n' in code_field_str:
                            code_field_array = [item.strip() for item in code_field_str.split('\n') if item.strip()]
                        else:
                            code_field_array = [code_field_str] if code_field_str else ["default_code_field"]
                    else:
                        code_field_array = ["default_code_field"]
                    
                    logger.info(f"Coincidencia por fallback encontrada. Raw code_field (col E): {code_field_raw}, Procesado como array: {code_field_array}")
                    return code_field_array
        
        logger.warning(f"No se encontró code_field para cuenta_contable '{cuenta_contable}', asignando valor por defecto")
        return [""]  # Devolver como array vacío
        
    except Exception as e:
        logger.error(f"Error obteniendo code_field para cuenta {cuenta_contable}: {str(e)}")
        return [""]  # Devolver como array vacío en caso de error

def obtener_item(libro, cuenta_contable, fila_encabezados=4):
    """
    Busca en la hoja 'Hoja5' la fila donde la columna de cuenta contable (columna B) coincida exactamente
    con el valor dado y retorna el valor de la columna 'Item' (columna C) de esa fila como string.
    Si no encuentra coincidencia, retorna una cadena vacía.
    """
    hoja = libro["Hoja5"]
    # Leer encabezados desde la fila indicada
    encabezados = [str(h).strip() if h is not None else "" for h in list(hoja.iter_rows(min_row=fila_encabezados, max_row=fila_encabezados, values_only=True))[0]]
    # Buscar el índice de la columna 'Item' (por defecto columna C = índice 2)
    try:
        idx_item = next(i for i, h in enumerate(encabezados) if h.strip().lower().replace(" ", "") == "item")
    except StopIteration:
        idx_item = 2  # Forzar columna C si no se encuentra el encabezado
    for fila in hoja.iter_rows(min_row=fila_encabezados+1, values_only=True):
        if fila[1] is not None:
            codigo_puc = str(fila[1]).strip()
            if codigo_puc == str(cuenta_contable).strip():
                valor_item = fila[idx_item] if idx_item < len(fila) else ""
                logger.info(f"Valor de 'Item' para cuenta_contable '{cuenta_contable}': {valor_item}")
                return str(valor_item) if valor_item is not None else ""
    logger.warning(f"No se encontró 'Item' para cuenta_contable '{cuenta_contable}' en Hoja5")
    return ""

# =============================
# Índices y limpieza en MongoDB
# =============================
def crear_indices_client_pucs(uid_usuario: str, ambiente):
    """
    Crea los índices necesarios para la colección client_pucs.
    """
    try:
        config = MongoDBConfig(env_prefix=ambiente)
        config.set_collection_name("client_pucs")
        gestor = MongoDBManager(config)
        gestor.collection.create_index(
            [("UID", 1), ("cuenta_contable", 1)],
            unique=True,
            name="uid_cuenta_contable_unique"
        )
        logger.info("Índice compuesto único creado: UID + cuenta_contable")
        gestor.close()
    except Exception as e:
        logger.error(f"Error creando índices: {str(e)}")
        logger.warning("Continuando sin crear índices...")

def eliminar_client_pucs_existentes(uid_usuario: str, ambiente):
    """
    Elimina todos los documentos existentes para el usuario en la colección client_pucs.
    """
    try:
        logger.info(f"DEBUG - Valor de ambiente recibido: '{ambiente}' (tipo: {type(ambiente)})")
        
        # Validar que el ambiente esté configurado
        if not ambiente or ambiente.upper() == "NONE":
            logger.error(f"Ambiente no válido: {ambiente}")
            sys.exit(1)
            
        logger.info(f"Configurando MongoDB con env_prefix: {ambiente}")
        config = MongoDBConfig(env_prefix=ambiente)
        
        # Debug: Mostrar la configuración generada
        logger.info(f"DEBUG - URI generada: {config.target_uri}")
        logger.info(f"DEBUG - DB name: {config.db_name}")
        
        config.set_collection_name("client_pucs")
        logger.info(f"DEBUG - Collection name: {config.get_collection_name()}")
        
        gestor = MongoDBManager(config)
        resultado = gestor.collection.delete_many({"UID": ObjectId(uid_usuario)})
        logger.info(f"Eliminados {resultado.deleted_count} documentos existentes para el usuario {uid_usuario}")
        gestor.close()
    except Exception as e:
        logger.error(f"Error eliminando documentos existentes: {str(e)}")
        sys.exit(1)

# =============================
# Procesamiento y subida de datos
# =============================
def procesar_archivo_excel(uid_usuario: str, ruta_xlsx: str, ambiente: str):
    """
    Procesa el archivo Excel y almacena los datos en MongoDB.
    """
    try:
        if not os.path.exists(ruta_xlsx):
            logger.error(f"Archivo Excel no encontrado: {ruta_xlsx}")
            sys.exit(1)
        logger.info("Eliminando documentos existentes...")
        eliminar_client_pucs_existentes(uid_usuario, ambiente)
        logger.info("Creando índices...")
        crear_indices_client_pucs(uid_usuario, ambiente)
        config = MongoDBConfig(env_prefix=ambiente)
        config.set_collection_name("client_pucs")
        gestor = MongoDBManager(config)
        logger.info(f"Cargando archivo Excel: {ruta_xlsx}")
        libro = openpyxl.load_workbook(ruta_xlsx, data_only=True)
        if "Hoja1" not in libro.sheetnames:
            logger.error(f"Hoja 'Hoja1' no encontrada en el archivo. Hojas disponibles: {libro.sheetnames}")
            sys.exit(1)
        encabezados = obtener_encabezados_excel(libro)
        hoja = libro["Hoja1"]
        filas_procesadas = 0
        filas_omitidas = 0
        for idx, fila in enumerate(hoja.iter_rows(min_row=6, values_only=True), start=6):
            idx_cuenta = encabezados.index("CUENTA CONTABLE   (OBLIGATORIO)")
            cuenta_contable = str(fila[idx_cuenta]).strip() if fila[idx_cuenta] is not None else ""
            idx_centro = encabezados.index("CENTRO DE COSTO")
            centro_costo = str(fila[idx_centro]).strip() if fila[idx_centro] is not None else ""
            idx_nit = encabezados.index("NIT")
            nit = str(fila[idx_nit]).strip() if fila[idx_nit] is not None else ""
            idx_subcentro = encabezados.index("SUBCENTRO DE COSTO")
            subcentro_costo = str(fila[idx_subcentro]).strip() if fila[idx_subcentro] is not None else ""
            if not cuenta_contable:
                continue
            doc_existente = gestor.collection.find_one({
                "UID": ObjectId(uid_usuario),
                "cuenta_contable": cuenta_contable
            })
            if doc_existente:
                logger.warning(f"Cuenta contable '{cuenta_contable}' ya existe, saltando...")
                filas_omitidas += 1
                continue
            puc = cuenta_contable[:6]
            
            # Obtener el code_field desde la Hoja5 del Excel
            code_field = obtener_code_field(libro, cuenta_contable)
            descripcion = obtener_item(libro, cuenta_contable)
            
            if buscar_embedding(puc, ambiente):
                documento = {
                    "UID": ObjectId(uid_usuario),
                    "cuenta_contable": cuenta_contable,
                    "description": descripcion,
                    "code_field": code_field,
                    "createdAt": datetime.now(),
                    "updatedAt": datetime.now()
                }
            else:
                documento = {
                    "UID": ObjectId(uid_usuario),
                    "cuenta_contable": cuenta_contable,
                    "description": descripcion,
                    "code_field": code_field,
                    "uniquePuc": True,
                    "createdAt": datetime.now(),
                    "updatedAt": datetime.now()
                }
            gestor.collection.insert_one(documento)
            crear_centro_costo_por_puc(uid_usuario, nit, cuenta_contable, centro_costo, subcentro_costo, ambiente)
            filas_procesadas += 1
            logger.info(f"Documento creado: {documento}")
            if filas_procesadas % 100 == 0:
                logger.info(f"Procesadas {filas_procesadas} filas...")
        logger.info(f"Proceso completado. Se procesaron {filas_procesadas} filas en total.")
        if filas_omitidas > 0:
            logger.info(f"Se saltaron {filas_omitidas} filas con cuentas contables duplicadas.")
        gestor.close()
    except Exception as e:
        logger.error(f"Error procesando archivo Excel: {str(e)}")
        sys.exit(1)

# =============================
# Utilidades de búsqueda y limpieza
# =============================
def buscar_embedding(puc: str, ambiente: str):
    """
    Busca si existe un embedding para el PUC dado en la colección puc_embeddings.
    """
    try:
        config = MongoDBConfig(env_prefix=ambiente)
        config.set_collection_name("puc_embeddings")
        gestor_embedding = MongoDBManager(config)
        existe_puc = gestor_embedding.collection.find_one({"code": puc})
        return bool(existe_puc)
    except Exception as e:
        logger.error(f"Error buscando embedding para el PUC {puc}: {str(e)}")
        return False

def limpiar_nit(nit):
    """
    Limpia el NIT/cédula de caracteres no numéricos y espacios.
    """
    if isinstance(nit, str):
        return ''.join(filter(str.isdigit, nit))
    return nit

# =============================
# Creación de centro de costo por PUC
# =============================
def crear_centro_costo_por_puc(uid: str, nit: str, cuenta: str, centro: str, subcentro: str = None, ambiente=None):
    """
    Crea un documento de centro de costo por PUC en la colección correspondiente.
    """
    logger.info(f"Iniciando creación de centro de costo para PUC: UID={uid}, NIT={nit}, Cuenta={cuenta}, CentroCosto={centro}, SubCentro={subcentro}")
    config = MongoDBConfig(env_prefix=ambiente)
    config.set_collection_name("cost_center_per_puc")
    gestor = MongoDBManager(config)
    nit_limpio = limpiar_nit(nit)
    centro_costo_obj = {"center": centro}
    if subcentro:
        centro_costo_obj["subcenter"] = subcentro
    doc_existente = gestor.collection.find_one({
        "UID": ObjectId(uid),
        "id_supplier": nit_limpio,
        "account_code": cuenta,
        "cost_center": centro_costo_obj,
    })
    if doc_existente:
        logger.warning(f"Ya existe un centro de costo para la cuenta '{cuenta}' y NIT '{nit_limpio}', se omite...")
        return
    documento = {
        "UID": ObjectId(uid),
        "id_supplier": nit_limpio,
        "account_code": cuenta,
        "cost_center": centro_costo_obj
    }
    resultado = gestor.collection.insert_one(documento)
    logger.info(f"Centro de costo por PUC creado con _id: {resultado.inserted_id} para NIT: {nit_limpio}, Cuenta: {cuenta}, CentroCosto: {centro_costo_obj}")

# =============================
# Función principal
# =============================
def main(uid_usuario=None, ruta_xlsx=None, ambiente=None):
    """
    Orquesta el proceso completo de onboarding de causación.
    Recibe el UID y la ruta del archivo Excel como argumentos.
    Si no se proporcionan, los toma de la línea de comandos o usa la ruta por defecto.
    """
    logger.info("Iniciando procesamiento del archivo de causación...")
    # Convertir uid_usuario a ObjectId si es necesario
    if uid_usuario is None:
        if len(sys.argv) < 3:
            print("Uso: python onboarding_causacion.py <UID> <AMBIENTE> [ruta_xlsx]")
            sys.exit(1)
        try:
            uid_usuario = ObjectId(sys.argv[1])
        except Exception:
            print("El UID proporcionado no es un ObjectId válido.")
            sys.exit(1)
        ambiente = sys.argv[2]
        if len(sys.argv) > 3:
            ruta_xlsx = sys.argv[3]
    else:
        if not isinstance(uid_usuario, ObjectId):
            try:
                uid_usuario = ObjectId(uid_usuario)
            except Exception:
                print("El UID proporcionado no es un ObjectId válido.")
                sys.exit(1)
    if ruta_xlsx is None:
        app_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
        ruta_xlsx = os.path.abspath(os.path.join(app_root, "data", "modelos_causacion", "SurtifloraModeloCausacionAbril2025.xlsx"))
    procesar_archivo_excel(str(uid_usuario), ruta_xlsx, ambiente)
    logger.info("Proceso completado exitosamente.")

if __name__ == "__main__":
    main()
